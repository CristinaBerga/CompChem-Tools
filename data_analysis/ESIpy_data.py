####################################################################
# Script: ESIpy_data.py
# Author: Cristina Berga, https://github.com/CristinaBerga/CompChem-Tools/
# Creation Date: August 2025
####################################################################
#
# This Python script is a powerful tool for the automated extraction
# and visualization of aromaticity indices from calculations performed
# with the ESI_py program (https://github.com/jgrebol/ESIpy).
#
####################################################################
#
# Key Features:
#
# 1. Custom Data Extraction: The script offers a flexible interface
#    to select specific aromaticity indices (e.g., HOMA, PDI, GEO)
#    and atomic partitions (e.g., MULLIKEN, LOWDIN) to be analyzed.
#
# 2. Automatically processes all .esi files found in the current
#    directory.
#
# 3. Structured Data Export: Consolidates the extracted data into a
#    single Excel file (`ESIpy_data.xlsx`). Each .esi file is
#    given its own sheet in the workbook.
#
# 4. Data Visualization: Maps the calculated aromaticity indices
#    directly onto the molecular geometry. It generates PNG images for
#    each molecule and each selected index, superimposing the numerical
#    values at the center of the corresponding aromatic rings.
#
####################################################################
#
# Usage:
#
# - Place the script in a directory containing your .esi and .out files.
#
# - Run the script from the command line. You will be prompted to
#   select which aromatic indices and atomic partitions you wish to
#   include in the analysis.
#
# - The final Excel file and all generated images will be saved in the same
#   directory.
#
####################################################################

import os
import re
import pandas as pd
from openpyxl import Workbook
from rdkit import Chem
from rdkit.Chem import AllChem
from rdkit.Chem.Draw import rdMolDraw2D
from rdkit.Geometry import Point2D
from PIL import Image, ImageDraw, ImageFont
import math
import io
import subprocess
import numpy as np

def get_user_choices(available_indices, available_partitions):
    """Obtain user preferences for indices and partitions."""
    print("Aromatic indices:")
    for idx, name in enumerate(available_indices, 1):
        print(f"  {idx}: {name}")
    print(f"  {len(available_indices) + 1}: All")

    index_choice = input(f"Enter the numbers of the indices you want (separated by commas) or '{len(available_indices) + 1}' for all: ")
    if index_choice == str(len(available_indices) + 1):
        selected_indices = available_indices
    else:
        try:
            chosen_indices = [int(x.strip()) for x in index_choice.split(',')]
            selected_indices = [available_indices[i - 1] for i in chosen_indices if 0 < i <= len(available_indices)]
        except (ValueError, IndexError):
            print("Invalid input. All indices will be used by default.")
            selected_indices = available_indices

    print("\nAtomic partitions:")
    for idx, name in enumerate(available_partitions, 1):
        print(f"  {idx}: {name}")
    print(f"  {len(available_partitions) + 1}: All")

    partition_choice = input(f"Enter the numbers of the partitions you want (separated by commas) or '{len(available_partitions) + 1}' for all: ")
    if partition_choice == str(len(available_partitions) + 1):
        selected_partitions = available_partitions
    else:
        try:
            chosen_partitions = [int(x.strip()) for x in partition_choice.split(',')]
            selected_partitions = [available_partitions[i - 1] for i in chosen_partitions if 0 < i <= len(available_partitions)]
        except (ValueError, IndexError):
            print("Invalid input. All partitions will be used by default.")
            selected_partitions = available_partitions
    
    return selected_indices, selected_partitions
    
def get_image_choice():
    """Ask the user if they want to generate images."""
    while True:
        choice = input("Do you want to generate molecule images with the indices? (y/n): ").lower().strip()
        if choice in ['yes', 'y']:
            return True
        elif choice in ['no', 'n']:
            return False
        else:
            print("Invalid response. Please enter 'yes' or 'no'.")

def get_mol_from_gaussian_out(filepath):
    """
    Parses a Gaussian .out file to get two molecule objects.
    """
    try:
        output_mol_path = 'temp_map.mol'
        subprocess.check_output(['obabel', filepath, '-O', output_mol_path], stderr=subprocess.STDOUT)
        mol_map = Chem.MolFromMolFile(output_mol_path, removeHs=False)
        os.remove(output_mol_path)
        if not mol_map:
            print(f"Could not create a mapping RDKit molecule from {filepath}.")
            return None, None
        Chem.SanitizeMol(mol_map)
        
        output_draw_mol_path = 'temp_draw.mol'
        subprocess.check_output(['obabel', filepath, '-O', output_draw_mol_path], stderr=subprocess.STDOUT) # Sometimes this keyword is necessary '--gen2d'
        mol_draw = Chem.MolFromMolFile(output_draw_mol_path, removeHs=False)
        os.remove(output_draw_mol_path)
        if not mol_draw:
            print(f"Could not create a drawing RDKit molecule from the 2D mol file for {filepath}.")
            return mol_map, None
        
        if mol_map.GetNumAtoms() == 1 or mol_draw.GetNumAtoms() == 1:
            return None, None

        return mol_map, mol_draw
        
    except FileNotFoundError:
        print("Open Babel (obabel) not found. Please make sure it's installed and in your PATH.")
        return None, None
    except subprocess.CalledProcessError as e:
        print(f"Error calling Open Babel for {filepath}: {e.stderr}")
        return None, None
    except Exception as e:
        print(f"Error parsing Gaussian .out file {filepath}: {e}")
        return None, None

def find_rings(mol):
    """
    Finds the rings in a molecule using RDKit's SSSR algorithm.
    Returns a list of lists with the atom indices of each ring.
    """
    try:
        rings = mol.GetRingInfo().AtomRings()
        return list(rings)
    except Exception as e:
        print(f"Error finding rings with RDKit: {e}")
        return []

def map_esi_to_rings(mol_map, ring_info_map, excel_data):
    """
    Maps the data from the Excel file to the correct rings in the molecule
    based on atom numbers.
    """
    excel_ring_map = {}
    for _, row in excel_data.iterrows():
        atom_numbers = set(int(x) - 1 for x in row['Atom numbers'].split())
        excel_ring_map[frozenset(atom_numbers)] = row
    
    ring_data_map = {}
    for i, rdkit_ring_indices in enumerate(ring_info_map):
        ring_set = frozenset(rdkit_ring_indices)
        if ring_set in excel_ring_map:
            ring_data_map[i] = excel_ring_map[ring_set]
        else:
            print(f"Warning: No matching data found in Excel for RDKit Ring {i+1} with atoms {list(ring_set)}.")
            
    return ring_data_map

def are_rings_identical(mol1, ring_atoms1, mol2, ring_atoms2):
    """
    Compares two rings from different molecules to see if they are topologically identical.
    """
    if len(ring_atoms1) != len(ring_atoms2):
        return False
    
    ring_types1 = sorted([mol1.GetAtomWithIdx(a).GetSymbol() for a in ring_atoms1])
    ring_types2 = sorted([mol2.GetAtomWithIdx(a).GetSymbol() for a in ring_atoms2])
    if ring_types1 != ring_types2:
        return False
        
    try:
        submol1 = Chem.RWMol()
        submol2 = Chem.RWMol()

        old_to_new_map1 = {atom_idx: submol1.AddAtom(mol1.GetAtomWithIdx(atom_idx)) for atom_idx in ring_atoms1}
        for bond in mol1.GetBonds():
            if bond.GetBeginAtomIdx() in old_to_new_map1 and bond.GetEndAtomIdx() in old_to_new_map1:
                submol1.AddBond(old_to_new_map1[bond.GetBeginAtomIdx()], old_to_new_map1[bond.GetEndAtomIdx()], bond.GetBondType())

        old_to_new_map2 = {atom_idx: submol2.AddAtom(mol2.GetAtomWithIdx(atom_idx)) for atom_idx in ring_atoms2}
        for bond in mol2.GetBonds():
            if bond.GetBeginAtomIdx() in old_to_new_map2 and bond.GetEndAtomIdx() in old_to_new_map2:
                submol2.AddBond(old_to_new_map2[bond.GetBeginAtomIdx()], old_to_new_map2[bond.GetEndAtomIdx()], bond.GetBondType())

        return submol1.GetMol().HasSubstructMatch(submol2.GetMol())

    except Exception as e:
        print(f"Error comparing rings: {e}")
        return False
    
def find_ring_mapping(mol_map, mol_draw):
    """
    Maps rings from the mapping molecule to the drawing molecule based on a
    more robust topological comparison.
    """
    map_rings = find_rings(mol_map)
    draw_rings = find_rings(mol_draw)
    mapping = {}
    
    mapped_draw_indices = set()
    for i, map_ring_atoms in enumerate(map_rings):
        for j, draw_ring_atoms in enumerate(draw_rings):
            if j not in mapped_draw_indices and are_rings_identical(mol_map, map_ring_atoms, mol_draw, draw_ring_atoms):
                mapping[i] = j
                mapped_draw_indices.add(j)
                break
    return mapping

def draw_molecule_with_indices(mol_draw, mol_map, ring_data_map, index, filename):
    """
    Dibuixa la molècula i col·loca les etiquetes centrades al mig de l'anell.
    - Utilitza mol_draw per a la visualització 2D correcta.
    - Utilitza mol_map i ring_data_map per a la correspondència de dades.
    """
    try:
        mol_no_h = Chem.RemoveHs(mol_draw)
        width, height = 1000, 1000
        drawer = rdMolDraw2D.MolDraw2DCairo(width, height)
        opts = drawer.drawOptions()
        opts.clearBackground = False
        rdMolDraw2D.PrepareAndDrawMolecule(drawer, mol_no_h)

        ring_info_map = find_rings(mol_map)
        ring_info_draw = find_rings(mol_draw)
        
        ring_mapping = find_ring_mapping(mol_map, mol_draw)

        centers_and_texts = []
        for i, map_ring_idx in enumerate(ring_info_map):
            if len(map_ring_idx) != 6:
                continue

            if i in ring_data_map and index in ring_data_map[i]:
                value = ring_data_map[i][index]
                if isinstance(value, float):
                    text = f"{value:.4f}"

                    if i in ring_mapping:
                        draw_ring_idx = ring_info_draw[ring_mapping[i]]
                        
                        conf = mol_draw.GetConformer()
                        xs = [conf.GetAtomPosition(k).x for k in draw_ring_idx]
                        ys = [conf.GetAtomPosition(k).y for k in draw_ring_idx]
                        cx, cy = sum(xs) / len(xs), sum(ys) / len(ys)
                        pt_draw = drawer.GetDrawCoords(Point2D(cx, cy))
                        centers_and_texts.append((pt_draw.x, pt_draw.y, text))
                    else:
                        print(f"Warning: Could not find a matching ring for drawing for ring with atoms {map_ring_idx}.")

        drawer.FinishDrawing()
        png_bytes = drawer.GetDrawingText()

        img = Image.open(io.BytesIO(png_bytes)).convert("RGBA")
        draw = ImageDraw.Draw(img)
        
        try:
            font = ImageFont.truetype("arial.ttf", 40)
        except IOError:
            font = ImageFont.load_default()

        pad = 6
        for x, y, text in centers_and_texts:
            bbox = draw.textbbox((0, 0), text, font=font)
            tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
            tx = x - tw / 2
            ty = y - th / 2
            #draw.rectangle([tx - pad, ty - pad, tx + tw + pad, ty + th + pad], fill="white", outline="withe")
            draw.text((tx, ty), text, fill=(51, 157, 235), font=font)

        img.save(filename)

    except Exception as e:
        print(f"Error generating image {filename}: {e}")

if __name__ == "__main__":
    index_patterns = {
        "PDI": r'\|\s*PDI\s+\s*\d+\s*=\s*([-]?\d+\.\d+)\s*',
        "MCI": r'\|\s*MCI\s+\s*\d+\s*=\s*([-]?\d+\.\d+)\s*',
        "MCI_NONM": r'\|\s*MCI\*\*\(1\/n\)\s+\d+\s*=\s*([-]?\d+\.\d+)\s*',
        "IRING": r'\|\s*Iring\s+\s*\d+\s*=\s*([-]?\d+\.\d+)\s*',
        "IRING_NORM": r'\|\s*Iring\*\*\(1\/n\)\s+\d+\s*=\s*([-]?\d+\.\d+)\s*',
        "EN": r'\|\s*EN\s+\s*\d+\s*=\s*([-]?\d+\.\d+)\s*',
        "GEO": r'\|\s*GEO\s+\s*\d+\s*=\s*([-]?\d+\.\d+)\s*',
        "HOMA": r'\|\s*HOMA\s+\s*\d+\s*=\s*([-]?\d+\.\d+)\s*',
        "FLU": r'\|\s*FLU\s+\s*\d+\s*=\s*([-]?\d+\.\d+)\s*',
        "BOA": r'\|\s*BOA\s+\s*\d+\s*=\s*([-]?\d+\.\d+)\s*',
        "BOA_cc": r'\|\s*BOA_cc\s+\s*\d+\s*=\s*([-]?\d+\.\d+)\s*',
    }
    available_indices = list(index_patterns.keys())
    available_partitions = ["MULLIKEN", "META_LOWDIN", "IAO", "LOWDIN", "NAO"]

    selected_indices, selected_partitions = get_user_choices(available_indices, available_partitions)
    if not selected_indices:
        print("No indices have been selected. The script has been stopped.")
        exit()

    generate_images = get_image_choice()

    headers = ["Atomic partition", "Ring"] + selected_indices + ["Atom numbers"]
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    esi_files = [f for f in os.listdir('.') if f.endswith(".esi")]
    if not esi_files:
        print("No .esi files found in the current directory.")
        exit()

    for filename in esi_files:
        print(f"Processing {filename}...")
        sheet_name = os.path.splitext(filename)[0][:31].replace(" ", "_")
        i = 1
        original_sheet_name = sheet_name
        while sheet_name in wb.sheetnames:
            sheet_name = f"{original_sheet_name[:28]}_{i}"
            i += 1
        
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)

        try:
            with open(filename, 'r', encoding='utf-8') as f:
                content = f.read()

            atomic_partition_sections = re.split(r'\|\s*Atomic partition:\s+(.*?)\s*\n', content)
            
            for i in range(1, len(atomic_partition_sections), 2):
                partition = atomic_partition_sections[i].strip()
                if selected_partitions and partition not in selected_partitions:
                    continue

                partition_content = atomic_partition_sections[i+1]
                rings = re.findall(r'\|\s*Ring\s+(\d+)\s+\((\d+)\):\s+([\d\s]+)', partition_content)
                
                for ring_number, ring_size, ring_atoms in rings:
                    row_data = [partition, f"Ring {ring_number} ({ring_size})"]

                    for index in selected_indices:
                        pattern = index_patterns[index].replace(r'\d+', ring_number, 1)
                        match = re.search(pattern, partition_content)
                        value = match.group(1) if match else 'N/A'
                        try:
                            row_data.append(float(value))
                        except (ValueError, TypeError):
                            row_data.append(value)
                    
                    row_data.append(ring_atoms.strip())
                    ws.append(row_data)
        
        except Exception as e:
            print(f"Error processing {filename}: {e}")

    output_filename = "ESIpy_data.xlsx"
    wb.save(output_filename)
    print(f"Excel file generated successfully: {output_filename}")
    
    if generate_images:
        print("\n--- Starting molecule image generation ---")
        
        for esi_filepath in esi_files:
            sheet_name = os.path.splitext(esi_filepath)[0][:31].replace(" ", "_")
            gaussian_out_filepath = os.path.splitext(esi_filepath)[0] + '.out'
            
            if not os.path.exists(gaussian_out_filepath):
                print(f"No corresponding Gaussian .out file found for {esi_filepath}. Skipping image generation.")
                continue
                
            mol_map, mol_draw = get_mol_from_gaussian_out(gaussian_out_filepath)
            
            if mol_map is None or mol_draw is None:
                print(f"Could not create valid RDKit molecules from {gaussian_out_filepath}. Skipping image generation.")
                continue
                
            ring_atom_groups = find_rings(mol_map)
            
            if not ring_atom_groups:
                print(f"No rings found in the molecule from {gaussian_out_filepath}. Skipping image generation.")
                continue
                
            try:
                df = pd.read_excel(output_filename, sheet_name=sheet_name)
            except Exception as e:
                print(f"Could not read sheet '{sheet_name}' from Excel file. Skipping image generation for {esi_filepath}.")
                continue
                
            for partition in selected_partitions:
                partition_df = df[df['Atomic partition'] == partition]
                
                ring_data_map = map_esi_to_rings(mol_map, ring_atom_groups, partition_df)
                
                for index in selected_indices:
                    if index not in partition_df.columns:
                        continue
                    
                    output_image_name = f"{os.path.splitext(esi_filepath)[0]}_{index}_{partition}.png".replace(" ", "_")
                    print(f"Generating image for: {output_image_name}")
                    draw_molecule_with_indices(mol_draw, mol_map, ring_data_map, index, output_image_name)

        print("\nAll molecule images have been successfully generated.")
